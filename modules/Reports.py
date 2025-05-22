from openpyxl import *
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font
from tkinter.filedialog import asksaveasfilename
import os

class Reports():
    def __init__(self):
        self.wb: Workbook
        self.sheets = {}        

    def createWorkbook(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.worksheets[0])

    def createSheet(self,name):
        sheet = self.wb.create_sheet(name)
        self.sheets.setdefault(name, sheet)
        return sheet

    def writeDataIn(self,sheetname, data, isBar):
        sheet = self.sheets[sheetname]
        for row in data:
            sheet.append(row)

        if isBar: chart = BarChart()
        else: chart = PieChart()
        categories = Reference(sheet, min_col=1, min_row=2, max_row=len(data))
        values = Reference(sheet, min_col=2, min_row=1, max_row=len(data))
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)

        sheet.add_chart(chart, "E1")

        columna = sheet['A']
        for celda in columna:
            celda.value = str(celda.value)

        column_letter = 'A'
        column_width = max(len(cell.value)*1.1 for cell in sheet[column_letter])
        sheet.column_dimensions[column_letter].width = column_width

        sheet.column_dimensions['B'].width = 15

        for c in ["A1","B1"]:
            cell = sheet[c]
            font = Font(bold=True) 
            cell.font = font

    def saveWorkbook(self): 
        archivo_guardado = asksaveasfilename(filetypes=(("Archivos de Excel", "*.xlsx;*.xls"),
                                                ("Todos los archivos", "*.*")),
                                            title="Guardar archivo como...")
        
        if archivo_guardado:
            carpeta = os.path.dirname(archivo_guardado)
            nombre_archivo = os.path.basename(archivo_guardado)

            ruta_archivo = os.path.join(carpeta, nombre_archivo)
            self.wb.save(f"{ruta_archivo}.xlsx")

            return True
        return False
